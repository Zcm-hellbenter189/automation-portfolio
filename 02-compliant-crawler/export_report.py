"""导出 Excel 报表 + 评分分布统计

用法: python export_report.py  （需先运行 python main.py 生成数据）
"""
import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
DB = BASE_DIR / "data" / "books.db"
OUT = BASE_DIR / "export" / "图书榜单报表.xlsx"


def load_records() -> list[dict]:
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    try:
        return [dict(r) for r in conn.execute(
            "SELECT * FROM books ORDER BY rating DESC, rating_count DESC")]
    finally:
        conn.close()


def build_report(records: list[dict]) -> Path:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # sheet 1: 榜单明细
    ws = wb.active
    ws.title = "榜单明细"
    headers = ["排名", "书名", "作者", "评分", "评价数", "一句话简介"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for i, r in enumerate(records, start=1):
        ws.append([i, r["title"], r["author"], r["rating"], r["rating_count"], r["quote"]])
    for col in range(1, len(headers) + 1):
        width = 40 if col in (2, 6) else 14
        ws.column_dimensions[get_column_letter(col)].width = width

    # sheet 2: 评分分布
    ws2 = wb.create_sheet("评分分布")
    buckets = {"9.0+": 0, "8.0-8.9": 0, "7.0-7.9": 0, "<7.0": 0}
    for r in records:
        s = r["rating"]
        if s >= 9.0:
            buckets["9.0+"] += 1
        elif s >= 8.0:
            buckets["8.0-8.9"] += 1
        elif s >= 7.0:
            buckets["7.0-7.9"] += 1
        else:
            buckets["<7.0"] += 1
    ws2.append(["评分区间", "数量"])
    ws2.append(["总计", len(records)])
    for k, v in buckets.items():
        ws2.append([k, v])

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    records = load_records()
    if not records:
        print("[!] 数据库为空，请先运行: python main.py")
    else:
        out = build_report(records)
        print(f"报表已生成: {out}（共 {len(records)} 条记录）")
