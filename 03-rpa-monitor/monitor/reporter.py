"""Excel 报表：概览 + 明细 + 告警 三个 sheet"""
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class Reporter:
    def __init__(self, report_dir: Path):
        self.report_dir = report_dir

    def build(self, samples: list[dict], alerts: list[dict]) -> Path:
        self.report_dir.mkdir(parents=True, exist_ok=True)
        out = self.report_dir / f"监控报表_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb = openpyxl.Workbook()

        # sheet 1: 监控概览
        ws = wb.active
        ws.title = "监控概览"
        ok_count = sum(1 for s in samples if not s.get("error"))
        avg_elapsed = round(sum(s["elapsed"] for s in samples) / max(len(samples), 1), 3)
        ws.append(["监控时间", time.strftime("%Y-%m-%d %H:%M:%S")])
        ws.append(["监控目标数", len(samples)])
        ws.append(["正常数", ok_count])
        ws.append(["异常数", len(alerts)])
        ws.append(["平均响应时间(s)", avg_elapsed])
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 30

        # sheet 2: 采样明细
        ws2 = wb.create_sheet("采样明细")
        headers = ["名称", "URL", "状态码", "响应时间(s)", "异常原因", "时间"]
        ws2.append(headers)
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws2.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for s in samples:
            ws2.append([s["name"], s["url"], s["status"], s["elapsed"],
                        s.get("error") or "", s["timestamp"]])
        for col in range(1, len(headers) + 1):
            ws2.column_dimensions[get_column_letter(col)].width = 45 if col == 2 else 18

        # sheet 3: 告警记录
        ws3 = wb.create_sheet("告警记录")
        ws3.append(["名称", "URL", "状态码", "响应时间(s)", "告警原因", "时间"])
        if alerts:
            for a in alerts:
                ws3.append([a["name"], a["url"], a["status"], a["elapsed"],
                            a["error"], a["timestamp"]])
        else:
            ws3.append(["本次监控无告警"])

        wb.save(out)
        return out
